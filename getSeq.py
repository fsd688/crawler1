# from selenium import webdriver
# browser = webdriver.Chrome(executable_path=r"C:\Program Files\Google\chromedriver.exe")
# # browser = webdriver.PhantomJS()
# browser.get("https://pubchem.ncbi.nlm.nih.gov/compound/12303645")
# a = browser.page_source
#
# print()

import pandas as pd
import requests
import openpyxl

workbook = openpyxl.load_workbook("./temp2.xlsx")
sheet = workbook["Sheet1"]
k = 1
j = 0
ids = pd.read_csv("./a.txt",header=None).values
for i in range(0,len(ids)):
    query = ids[i][0]
    url = "https://rest.uniprot.org/uniprotkb/"+query+".fasta"
    request = requests.get(url)
    html_doc = request.text
    print(query,html_doc[html_doc.find('\n')+1:])
    sheet.cell(row=k, column=1, value=query)
    sheet.cell(row=k, column=2, value=html_doc[html_doc.find('\n')+1:])
    k+=1

workbook.save('./temp2.xlsx')






