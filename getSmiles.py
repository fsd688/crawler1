import requests
from bs4 import BeautifulSoup
import xlrd
import xlwt
import openpyxl
import pubchempy as pcp
from selenium import webdriver
import time
import re
# 读取中药成分
workbook = openpyxl.load_workbook('./成分特征.xlsx')
sheet = workbook["Sheet2"]
i=j=1
with xlrd.open_workbook('./成分特征.xlsx') as data_excel:
    names = data_excel.sheet_names()
    table = data_excel.sheet_by_name(sheet_name='Sheet1')
    # table2 = data_excel.sheet_by_name(sheet_name="爬虫（成分对应smiles）")
    table_rows = table.nrows
    mol_list = table.col(colx=0)[0:][:]
    for mol_name in mol_list:
        query_number =int(mol_name.value[3:])
        url = 'https://old.tcmsp-e.com/molecule.php?qn='+str(query_number)
        request = requests.get(url)
        html_doc = request.text
        soup = BeautifulSoup(html_doc,"html.parser")
        pubchem_cid = soup.find_all("td")[5].text
        if(pubchem_cid!='N/A'):
            # pubchem_url = 'https://pubchem.ncbi.nlm.nih.gov/compound/'+pubchem_cid+"/JSON/"
            # request = requests.get(pubchem_url)
            # html_doc = request.text
            # soup = BeautifulSoup(html_doc, "html.parser")
            # print()
            # c = pcp.Compound.from_cid(pubchem_cid)
            # print(mol_name.value,c.canonical_smiles)
            # sheet.cell(row=i, column=1, value=mol_name.value)
            # sheet.cell(row=i, column=2, value=c.canonical_smiles)
            print(mol_name.value,'c.canonical_smiles')
            i+=1
        elif(pubchem_cid=='N/A'):
            print(mol_name.value, "n/a")
workbook.save("./temp.xlsx")








